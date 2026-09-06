"""Local, bounded document extraction. PDF parsing is owned by the native helper.

Office libraries read saved document content after bounded archive validation;
they never run Office, formulas, macros, links, or conversion programs. This
module also runs in a disposable child process so the job owner can enforce a
hard deadline and cancellation for every format.
"""
from __future__ import annotations

import csv
import hashlib
import json
import sys
import zipfile
from pathlib import Path
from typing import Any

FORMATS = {"pdf", "docx", "xlsx", "csv", "tsv"}
MAX_SOURCE_BYTES = 100 * 1024 * 1024
MAX_TEXT_BYTES = 5 * 1024 * 1024
MAX_PAGES = 500
MAX_CELLS = 200_000
EXTRACTOR_VERSION = "locus-documents-1"
MAX_XML_BYTES = 100 * 1024 * 1024


class ExtractionError(ValueError):
    pass


class _LimitReached(Exception):
    pass


class Segments:
    def __init__(self) -> None:
        self.items: list[dict[str, Any]] = []
        self.bytes = 0
        self.truncated = False
        self.warnings: list[str] = []

    def append(self, text: str, locator: dict[str, Any], method: str = "embedded") -> None:
        text = text.strip()
        if not text:
            return
        data = text.encode("utf-8")
        available = MAX_TEXT_BYTES - self.bytes
        if len(data) > available:
            text = data[:available].decode("utf-8", errors="ignore")
            self.truncated = True
            self.warnings.append("Text extraction reached the 5 MB limit.")
        if text:
            self.items.append({"text": text, "locator": locator, "method": method})
            self.bytes += len(text.encode("utf-8"))
        if self.truncated:
            raise _LimitReached

    def result(self) -> dict[str, Any]:
        return {
            "segments": self.items, "truncated": self.truncated,
            "warnings": self.warnings, "extractor_version": EXTRACTOR_VERSION,
        }


def _safe_zip(path: Path) -> zipfile.ZipFile:
    try:
        archive = zipfile.ZipFile(path)
        entries = archive.infolist()
        if len(entries) > 20_000 or sum(item.file_size for item in entries) > MAX_XML_BYTES:
            archive.close()
            raise ExtractionError("Expanded document exceeds the 100 MB safety limit.")
        if any(item.flag_bits & 1 for item in entries):
            archive.close()
            raise ExtractionError("Encrypted Office documents are not supported.")
        if len({item.filename for item in entries}) != len(entries):
            archive.close()
            raise ExtractionError("Document contains duplicate archive entries.")
        return archive
    except (zipfile.BadZipFile, OSError) as exc:
        raise ExtractionError("The Office document is damaged or unreadable.") from exc


def _docx(path: Path, out: Segments) -> None:
    from docx import Document
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    with _safe_zip(path) as archive:
        _preflight_xml(archive)
        has_notes = any(name in archive.namelist() for name in ("word/footnotes.xml", "word/endnotes.xml"))
    document = Document(path)
    paragraph = 0
    heading = ""

    def blocks(container: Any) -> None:
        nonlocal paragraph, heading
        for block in container.iter_inner_content():
            if isinstance(block, Paragraph):
                paragraph += 1
                if block.style is not None and "heading" in block.style.name.lower():
                    heading = block.text[:300]
                locator: dict[str, Any] = {"kind": "paragraph", "paragraph_start": paragraph, "paragraph_end": paragraph}
                if heading:
                    locator["heading"] = heading
                out.append(block.text, locator)
            elif isinstance(block, Table):
                seen_cells: set[Any] = set()
                for row in block.rows:
                    for cell in row.cells:
                        # Merged cells are repeated by the library's grid view.
                        if cell._tc not in seen_cells:
                            seen_cells.add(cell._tc)
                            blocks(cell)

    blocks(document)
    if has_notes:
        out.warnings.append("Footnotes and endnotes are not included in the body-text index.")


def _preflight_xml(archive: zipfile.ZipFile) -> None:
    for item in archive.infolist():
        if item.filename.lower().endswith((".xml", ".rels")):
            data = archive.read(item)
            if b"<!DOCTYPE" in data.upper() or b"<!ENTITY" in data.upper():
                raise ExtractionError("Document XML contains unsupported entity declarations.")


def _column(index: int) -> str:
    value = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        value = chr(65 + remainder) + value
    return value


def _xlsx(path: Path, out: Segments) -> None:
    from openpyxl import load_workbook

    cells_seen = 0
    with _safe_zip(path) as archive:
        _preflight_xml(archive)
    # Read-only streaming avoids materializing a giant worksheet; data_only
    # reads saved formula values, and keep_links=False never follows links.
    handle = path.open("rb")
    workbook = load_workbook(handle, read_only=True, data_only=True, keep_links=False)
    try:
        for sheet in workbook.worksheets:
            if sheet.sheet_state != "visible":
                out.warnings.append(f"Hidden sheet omitted: {sheet.title}")
                continue
            # Bounds include sparse/blank cells so forged dimensions cannot
            # force billions of iterations before the cell budget is checked.
            for row in sheet.iter_rows():
                values: list[str] = []
                refs: list[str] = []
                for cell in row:
                    cells_seen += 1
                    if cells_seen > MAX_CELLS:
                        out.truncated = True
                        out.warnings.append("Workbook reached the 200,000 cell limit.")
                        raise _LimitReached
                    if cell.value is None:
                        continue
                    refs.append(cell.coordinate)
                    values.append(f"{cell.coordinate}: {cell.value}")
                if refs:
                    out.append(" | ".join(values), {
                        "kind": "sheet", "sheet": sheet.title,
                        "cell_range": f"{refs[0]}:{refs[-1]}",
                    })
        out.warnings.append("Formula cells use their saved values; formulas are never calculated. Cells without a saved value are omitted.")
    finally:
        workbook.close()
        handle.close()


def _delimited(path: Path, out: Segments, delimiter: str) -> None:
    cells = 0
    # Decode strictly: silently replacing bytes can change a fact in a citation.
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle, delimiter=delimiter)
            for row_number, row in enumerate(reader, 1):
                cells += len(row)
                if cells > MAX_CELLS:
                    out.truncated = True
                    out.warnings.append("Table reached the 200,000 cell limit.")
                    raise _LimitReached
                if row:
                    out.append(" | ".join(f"{_column(i)}{row_number}: {value}" for i, value in enumerate(row, 1)), {
                        "kind": "sheet", "sheet": path.stem,
                        "cell_range": f"A{row_number}:{_column(len(row))}{row_number}",
                    })
    except (UnicodeError, csv.Error) as exc:
        raise ExtractionError("The table must be a valid UTF-8 CSV or TSV file.") from exc


def extract_office(path: Path, format: str, filename: str = "") -> dict[str, Any]:
    if path.stat().st_size > MAX_SOURCE_BYTES:
        raise ExtractionError("Document exceeds the 100 MB limit.")
    out = Segments()
    try:
        if format == "docx":
            _docx(path, out)
        elif format == "xlsx":
            _xlsx(path, out)
        elif format in {"csv", "tsv"}:
            _delimited(path, out, "\t" if format == "tsv" else ",")
        else:
            raise ExtractionError("Unsupported document format.")
    except _LimitReached:
        pass
    if not out.items and not out.truncated:
        out.warnings.append("No searchable text was found.")
    if format in {"csv", "tsv"} and filename:
        for segment in out.items:
            segment["locator"]["sheet"] = Path(filename).stem
    return out.result()


def main() -> None:
    """Private stdio protocol; stdout is exclusively versioned JSON records."""
    try:
        request = json.loads(sys.stdin.buffer.readline(65_537))
        if request.get("protocol_version") != 1:
            raise ExtractionError("Unsupported extraction protocol.")
        path = Path(str(request["path"]))
        expected = str(request.get("expected_hash") or "")
        digest = hashlib.sha256()
        with path.open("rb") as source:
            for chunk in iter(lambda: source.read(1024 * 1024), b""):
                digest.update(chunk)
        if not expected or digest.hexdigest() != expected:
            raise ExtractionError("Document snapshot changed before extraction.")
        result = extract_office(path, str(request["format"]), str(request.get("filename") or ""))
        for segment in result.pop("segments"):
            print(json.dumps({"type": "segment", "protocol_version": 1, **segment}), flush=True)
        print(json.dumps({"type": "result", "protocol_version": 1, "ok": True, **result}), flush=True)
    except Exception as exc:
        print(json.dumps({"type": "result", "protocol_version": 1, "ok": False, "error": str(exc)[:1000]}), flush=True)
        raise SystemExit(1) from None


if __name__ == "__main__":
    main()
